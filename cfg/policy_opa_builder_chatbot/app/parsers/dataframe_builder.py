"""Build pandas DataFrames and Excel files for YANG-derived rows."""

# dataframe_builder.py
from __future__ import annotations
import os
from typing import Iterable, List, Tuple, Optional, Union

import pandas as pd
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
import config.config as config
from parsers.utils import Row, get_logger, normalize_description, preserve_order, safe_filename

log = get_logger(__name__)

# The 8 column names (business schema) – keep header spelling stable
COLUMNS: Tuple[str, ...] = (
    "ManagedObject",
    "HierarchicalPath",
    "ParameterShortName",
    "ParameterLongName",
    "ParameterDescription",
    "DefaultValue",
    "Units",
    "Range",
    "Type",
    "isKey",
    "MinInstanceCardinality",
    "MaxInstanceCardinality"
)


def _row_key(r: Row) -> Tuple[str, str, str]:
    """De-duplication key: (MO, Path, ParamShort)."""
    return (
        r.managed_object or "",
        r.hierarchical_path or "",
        r.param_short or "",
    )


def _row_to_dict(r: Row) -> dict:  # type: ignore[type-arg]
    """Convert a Row dataclass to a flat dict matching COLUMNS."""
    return {
        "ManagedObject": r.managed_object or "",
        "HierarchicalPath": (r.hierarchical_path or "")
        # .replace("attributes", "")
        .replace("//", "/")
        .strip("/"),
        "ParameterShortName": r.param_short or "",
        "ParameterLongName": r.param_long or r.param_short or "",
        "ParameterDescription": normalize_description(r.description or ""),
        "DefaultValue": r.default or "",
        "Units": r.units or "",
        "Range": r.range_text or "",
        "Type": r.type_text or "",
        "isKey": (r.is_key or "").lower(),
        "MinInstanceCardinality": r.min_instance_cardinality or "",
        "MaxInstanceCardinality": r.max_instance_cardinality or "",
    }


class CsvWriter:
    """Facade for building DataFrames and writing Excel files.

    - `rows_to_dataframe(...)` returns a de-duplicated pandas DataFrame (8 columns).
    - `write(...)` writes an Excel (.xlsx) with frozen headers, auto widths, and wrapped text.
    """

    def __init__(
        self,
        out_dir: str,
        *,
        encoding: Optional[str] = None,  # kept for API compatibility (unused for Excel)
        delimiter: Optional[str] = None,
    ) -> None:  # kept for API compatibility
        """Initialize the writer with an output directory and CSV-like options."""
        self.out_dir = os.path.abspath(out_dir)
        self.encoding = encoding or config.CSV_ENCODING
        self.delimiter = delimiter or config.CSV_DELIMITER
        base_dir = Path(self.out_dir)
        base_dir.mkdir(parents=True, exist_ok=True)
        # os.makedirs(self.out_dir, exist_ok=True)

    def _xlsx_path_for(self, module_name: str) -> str:
        base = safe_filename(module_name or "module")
        return os.path.join(self.out_dir, f"{base}.xlsx")

    # ---------- public helper to build a DataFrame ----------
    def rows_to_dataframe(self, rows: Iterable[Row]) -> pd.DataFrame:
        """Convert Row items to a de-duplicated DataFrame.

        The resulting DataFrame uses the canonical 8 business columns in the
        prescribed order, with first-wins de-duplication by (MO, Path, ParamShort).
        """
        unique: List[Row] = preserve_order(rows, key=_row_key)
        dict_rows = [_row_to_dict(r) for r in unique]
        df = pd.DataFrame(dict_rows, columns=list(COLUMNS))
        return df

    # ----------wrap description---------------------
    def _wrap_description(self, ws: Worksheet, col_idx: Union[str, int]) -> None:
        """Wrap text in description column."""
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            row[0].alignment = Alignment(wrap_text=True, vertical="top")

    # ---------- write Excel (replaces CSV) ----------
    def write(self, module_name: str, rows: Iterable[Row]) -> str:
        """Write a per-module Excel file (.xlsx) with formatting.

        Produces a sheet with the 8 business columns, frozen header row, an
        autofilter, auto column widths, and wrapped text in the description column.
        Returns the full path to the written file.
        """
        df = self.rows_to_dataframe(rows)
        path = module_name
        # 1) Write DataFrame to Excel (Parameters sheet)
        with pd.ExcelWriter(path, engine="openpyxl") as xw:
            df.to_excel(xw, sheet_name="Parameters", index=False)

            # 2) Polish with openpyxl
            ws = xw.book["Parameters"]

            # Freeze header row
            ws.freeze_panes = "A2"

            # Autofilter across the header row
            ws.auto_filter.ref = ws.dimensions  # entire used range

            # Header styling: bold
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                cell.font = Font(bold=True)

            # Wrap text for the description column
            # Find the index of "ParameterDescription"
            desc_col_idx = (
                df.columns.get_loc("ParameterDescription") + 1
                if "ParameterDescription" in df.columns
                else None
            )  # 1-based
            # Auto column widths (rough heuristic; cap for very long text)
            MAX_COL_WIDTH = 80
            for col_idx, col_name in enumerate(df.columns, start=1):
                # Compute width based on header and sample of cell values
                header_len = len(str(col_name))
                sample_values = (
                    df[col_name].astype(str).tolist() if col_name in df.columns else []
                )
                max_len = (
                    max([header_len] + [len(str(v)) for v in sample_values])
                    if sample_values
                    else header_len
                )
                width = min(MAX_COL_WIDTH, max(12, max_len + 2))
                ws.column_dimensions[get_column_letter(col_idx)].width = width

                # Wrap description column cells
                if desc_col_idx == col_idx:
                    self._wrap_description(ws, col_idx)

        log.info("Wrote %s (%d rows)", path, len(df))
        return path


# Alias preserved for ergonomics
DataFrameBuilder = CsvWriter
